import numpy as np
from astropy.io import fits
#from astropy.wcs import WCS
from astropy import wcs   # A slight variation...
import tSZ_spectrum      as tsz
import kSZ_spectrum      as ksz
import astropy.units     as u
import my_astro_defaults as mad
import astropy.constants as const
import mapping_modules   as mm
from astropy.wcs import WCS
import ellipsoidal_shells as es
import image_filtering as imf
#import radec_image as radec
import importlib
import openpyxl

############################################################################
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
            
def inst_params(instrument):

    if instrument == "MUSTANG":
        fwhm1 = 8.7*u.arcsec  # arcseconds
        norm1 = 0.94          # normalization
        fwhm2 = 28.4*u.arcsec # arcseconds
        norm2 = 0.06          # normalization
        fwhm  = 9.0*u.arcsec
        smfw  = 10.0*u.arcsec
        freq  = 90.0*u.gigahertz # GHz
        FoV   = 42.0*u.arcsec #
        
    if instrument == "MUSTANG2":
        fwhm1 = 8.9*u.arcsec  # arcseconds
        norm1 = 0.97          # normalization
        fwhm2 = 25.0*u.arcsec # arcseconds
        norm2 = 0.03          # normalization
        fwhm  = 9.0*u.arcsec
        smfw  = 10.0*u.arcsec
        freq  = 90.0*u.gigahertz # GHz
        FoV   = 4.25*u.arcmin 
        
    if instrument == "NIKA":
        fwhm1 = 8.7*2.0*u.arcsec  # arcseconds
        norm1 = 0.94     # normalization
        fwhm2 = 28.4*2.0*u.arcsec # arcseconds
        norm2 = 0.06     # normalization
        fwhm  = 18.0*u.arcsec
        smfw  = 10.0*u.arcsec
        freq  = 150.0*u.gigahertz    # GHz
        FoV   = 2.15*u.arcmin 
        
    if instrument == "NIKA2":
        fwhm1 = 8.7*2.0*u.arcsec  # arcseconds
        norm1 = 0.94     # normalization
        fwhm2 = 28.4*2.0*u.arcsec # arcseconds
        norm2 = 0.06     # normalization
        fwhm  = 18.0*u.arcsec
        smfw  = 10.0*u.arcsec
        freq  = 150.0*u.gigahertz    # GHz
        FoV   = 6.5*u.arcmin 
        
    if instrument == "BOLOCAM":
        fwhm1 = 8.7*7.0*u.arcsec  # arcseconds
        norm1 = 0.94     # normalization
        fwhm2 = 28.4*7.0*u.arcsec # arcseconds
        norm2 = 0.06     # normalization
        fwhm  = 58.0*u.arcsec
        smfw  = 60.0*u.arcsec
        freq  = 140.0*u.gigahertz    # GHz
        FoV   = 8.0*u.arcmin #* (u.arcmin).to("arcsec")

    if instrument == "SZA":
        fwhm1 = 1.6*60.0*u.arcsec  # arcseconds
        norm1 = 1.0     # normalization
        fwhm2 = 1.0*u.arcsec # arcseconds
        norm2 = 0.00     # normalization
        fwhm  = 1.6*60.0*u.arcsec
        smfw  = 1.5*60.0*u.arcsec
        freq  = 31.0*u.gigahertz    # GHz
        FoV   = 30.0*u.arcmin #* (u.arcmin).to("arcsec")
        
    if instrument == "ACT90":
        fwhm1 = 2.16*60.0*u.arcsec  # arcseconds
        norm1 = 1.0     # normalization
        fwhm2 = 1.0*u.arcsec # arcseconds
        norm2 = 0.00     # normalization
        fwhm  = 2.16*60.0*u.arcsec
        smfw  = 2.0*60.0*u.arcsec
        freq  = 97.0*u.gigahertz    # GHz
        FoV   = 60.0*u.arcmin #* (u.arcmin).to("arcsec")
        
    if instrument == "ACT150":
        fwhm1 = 1.3*60.0*u.arcsec  # arcseconds
        norm1 = 1.0     # normalization
        fwhm2 = 1.0*u.arcsec # arcseconds
        norm2 = 0.00     # normalization
        fwhm  = 1.3*60.0*u.arcsec
        smfw  = 1.2*60.0*u.arcsec
        freq  = 148.0*u.gigahertz    # GHz
        FoV   = 60.0*u.arcmin #* (u.arcmin).to("arcsec")
        
    #mybv = get_bv_from_Jy2K(Jy2K)
#    else:
#        fwhm1=9.0*u.arcsec ; norm1=1.0
#        fwhm2=30.0*u.arcsec ; norm2=0.0
#        fwhm = 9.0*u.arcsec ; smfw = 10.0*u.arcsec
#        freq = 90.0*u.gigahertz 
#        FoV   = 1.0*u.arcmin * (u.arcmin).to("arcsec")
#        
#    import pdb; pdb.set_trace()

    return fwhm1,norm1,fwhm2,norm2,fwhm,smfw,freq,FoV

def inst_bp(instrument,array="2",addruze=False):
    """
    Returns a frequency and bandpass array
    
    Parameters
    __________
    instrument : MUSTANG, MUSTANG2, BOLOCAM, NIKA, or NIKA2
    currently only MUSTANG2 and NIKA2 are supported

    Returns
    -------
    -> farr   - The frequency array (in GHz)
    -> band   - the bandpass, with aperture efficiency applied
    """

    if instrument == "MUSTANG2" or instrument == "MUSTANG":
        srms = (300*u.um).to("m")        # surface RMS (microns)
        ### Reference: https://science.nrao.edu/facilities/gbt/proposing/GBTpg.pdf
        EA90 = 0.36   # Aperture efficiency at 90 GHz
        ### The beam efficiencies should be taken as 1.37* Aperture Efficiency
        R90  = np.exp(-4.0*np.pi*(srms/(const.c/(9.0e10*u.s**-1))).value)    #
        Gnot = EA90/R90                   # Unphysical, but see documentation...
        if instrument == "MUSTANG2":
            flow = 75.0   # GHz
            fhig = 105.0  # GHz
        else:
            flow = 82.5   # GHz
            fhig = 97.5  # GHz
            
        farr = np.arange(flow,fhig,1.0)   # frequency array.
        tran = np.ones(farr.shape)        # Let the transmission be unity everywhere.
        Larr = const.c.value/(farr*1.0e9) # Keep calm and carry on.

        tran = read_M2_bp(farr)           # New...as of 25 March 2020 ... OMG
        #import pdb;pdb.set_trace()
        
        ### Old formula:
        #Ruze = Gnot * np.exp(-4.0*np.pi*(srms.value)/Larr)
        ### Correct formula: (10 April 2018)
        Ruze = Gnot * np.exp(-(4.0*np.pi*srms.value/Larr)**2)
        NRuz = Ruze / np.max(Ruze)        # Normalize it
        band = tran * Ruze if addruze else tran # Bandpass, with (unnormalized) Ruze efficiency
       
    if instrument == "NIKA2" or instrument == "NIKA":
        caldir='/home/romero/NIKA2/NIKA_SVN/Processing/Pipeline/Calibration/BP/'
        bpfile=caldir+'Transmission_2017_Jan_NIKA2_v1.fits'
        hdulist = fits.open(bpfile)

        if array == "1H":      # 1mm (260 GHz) array, Horizontal Polarization
            tbdata = hdulist[1].data # 1H
            freq = tbdata.field(0)
            tran = tbdata.field(1)
            erro = tbdata.field(2)
            atmt = tbdata.field(3)
            cfreq1h = np.sum(freq*tran)/np.sum(tran)
        
        if array == "1V":     # 1mm (260 GHz) array, Vertical Polarization
            tbdata = hdulist[2].data # 1V
            freq = tbdata.field(0)
            tran = tbdata.field(1)
            erro = tbdata.field(2)
            atmt = tbdata.field(3)
            cfreq1v = np.sum(freq*tran)/np.sum(tran)
        
        if array == "2":       # 2mm (150 GHz) array
            tbdata = hdulist[3].data # 2
            freq = tbdata.field(0)
            tran = tbdata.field(1)
            erro = tbdata.field(2)
            atmt = tbdata.field(3)
            cfreq2 = np.sum(freq*tran)/np.sum(tran)

        ### Trim the zero-frequency listing, if any.
        gi=np.where(freq > 0)
        freq = freq[gi]
        tran = tran[gi]
        erro = erro[gi]
        atmt = atmt[gi]
        
### Calculate Aperture efficiencies from information found at:
### http://www.iram.es/IRAMES/mainwiki/Iram30mEfficiencies
        Beff = 0.630         # at 210 GHz
        Aeff = Beff/1.27     # See text on webpage
        srms = (66.0*u.um).to("m")        # surface RMS (microns)
        R210 = np.exp(-4.0*np.pi*(srms/(const.c/(2.1e11*u.s**-1))).value)    #
        Gnot = Aeff/R210                   # Unphysical, but see documentation...

        Larr = const.c.value/(freq*1.0e9) # Keep calm and carry on.        
        Ruze = Gnot * np.exp(-4.0*np.pi*(srms.value)/Larr)
        NRuz = Ruze / np.max(Ruze)        # Normalize it
        band = tran * Ruze if addruze else tran  # Bandpass, with (unnormalized) Ruze efficiency
        farr = freq
        
#########################################################################

    if instrument == 'ACT90':
        srms = (27.0*u.um).to("m")        # surface RMS (microns)
        EA90 = 0.95   # I'm making this number up...
        R90  = np.exp(-4.0*np.pi*(srms/(const.c/(9.0e10*u.s**-1))).value)    #
        Gnot = EA90/R90                   # Unphysical, but see documentation...
        flow = 65.0   # GHz
        fhig = 125.0  # GHz
        farr = np.arange(flow,fhig,1.0)  # frequency array.
        freq_ref = 90.0 # I took EA90 to be a fictitious aperature efficiency at 90 GHz
        band =  ruze_eff(farr,freq_ref,EA90,srms) if addruze else farr*0+1.0

    if instrument == 'ACT150':
        srms = (27.0*u.um).to("m")        # surface RMS (microns)
        EA90 = 0.95   #  I'm making this number up...
        R90  = np.exp(-4.0*np.pi*(srms/(const.c/(9.0e10*u.s**-1))).value)    #
        Gnot = EA90/R90                   # Unphysical, but see documentation...
        flow = 120.0   # GHz
        fhig = 180.0  # GHz
        farr = np.arange(flow,fhig,1.0)  # frequency array.
        freq_ref = 90.0 # I took EA90 to be a fictitious aperature efficiency at 90 GHz
        band =  ruze_eff(farr,freq_ref,EA90,srms) if addruze else farr*0+1.0


    return band, farr

def ruze_eff(freqs,freq_ref,ref_eff,srms):
    """
    freqs     - an array of frequencies in GHz (unitless within Python)
    freq_ref  - reference frequency     in GHz
    ref_eff   - reference aperature efficiency at the reference frequency.
    srms      - Surface RMS, in meters (with units in Python)
    """

    R_ref  = np.exp(-4.0*np.pi*(srms/(const.c/(freq_ref*1.0e9*u.s**-1))).value)    #
    Gnot   = ref_eff / R_ref
    
    tran = freqs*0.0 + 1.0            # Let the transmission be unity everywhere.
    Larr = const.c.value/(freqs*1.0e9) # Keep calm and carry on.
    ### Old formula:
    #Ruze = Gnot * np.exp(-4.0*np.pi*(srms.value)/Larr)
    ### Correct formula: (10 April 2018)
    Ruze = Gnot * np.exp(-(4.0*np.pi*srms.value/Larr)**2)
    NRuz = Ruze / np.max(Ruze)        # Normalize it
    band = tran * Ruze                # Bandpass, with (unnormalized) Ruze efficiency

    return band

def read_M2_bp(farr):

    m2bpfile='/home/data/MUSTANG2/InstrumentPerformance/BandPass/Filter_data.xlsm'
    mysheets=['K1860_2','K1887_2','K1888_2','K1889_2']
    # Lesson: Python will read in the **formula**, not the number (value)!!!
    # Fixed now.
    wb=openpyxl.load_workbook(m2bpfile)
    IR50  = 0.985  # Scalar transmission
    IR100 = 0.985  # Scalar transmission
    PTFE  = 0.99   # Scalar transmission
    Nylon = 0.27   # Scalar transmission

    trans = np.ones(farr.shape)
    for sname in mysheets:
        sheet=wb.get_sheet_by_name(sname)
        myfreq = []; mytran = []
        mymaxrow = sheet.max_row
        for row in range(3,100):
            myfreq.append(sheet.cell(row=row,column=1).value)
            mytran.append(sheet.cell(row=row,column=2).value)
        #print(farr[0],myfreq[0])
        #import pdb;pdb.set_trace()
        mytrans = np.interp(farr,myfreq,mytran)
        trans *= mytrans

    trans *= IR50*IR100*PTFE*Nylon
        
    return trans
    
def old_inst_bp(instrument,array="2"):
    """
    Returns a frequency and bandpass array
    
    Parameters
    __________
    instrument : MUSTANG, MUSTANG2, BOLOCAM, NIKA, or NIKA2
    currently only MUSTANG2 and NIKA2 are supported

    Returns
    -------
    -> farr   - The frequency array (in GHz)
    -> band   - the bandpass, with aperture efficiency applied
    """

    if instrument == "MUSTANG2" or instrument == "MUSTANG":
        srms = (300*u.um).to("m")        # surface RMS (microns)
        ### Reference: https://science.nrao.edu/facilities/gbt/proposing/GBTpg.pdf
        EA90 = 0.36   # Aperture efficiency at 90 GHz
        ### The beam efficiencies should be taken as 1.37* Aperture Efficiency
        R90  = np.exp(-4.0*np.pi*(srms/(const.c/(9.0e10*u.s**-1))).value)    #
        Gnot = EA90/R90                   # Unphysical, but see documentation...
        if instrument == "MUSTANG2":
            flow = 75.0   # GHz
            fhig = 105.0  # GHz
        else:
            flow = 82.5   # GHz
            fhig = 97.5  # GHz
            
        farr = np.arange(flow,fhig,1.0)  # frequency array.
        tran = farr*0.0 + 1.0            # Let the transmission be unity everywhere.
        Larr = const.c.value/(farr*1.0e9) # Keep calm and carry on.
        ### Old formula:
        #Ruze = Gnot * np.exp(-4.0*np.pi*(srms.value)/Larr)
        ### Correct formula: (10 April 2018)
        Ruze = Gnot * np.exp(-(4.0*np.pi*srms.value/Larr)**2)
        NRuz = Ruze / np.max(Ruze)        # Normalize it
        band = tran * Ruze                # Bandpass, with (unnormalized) Ruze efficiency
       
    if instrument == "NIKA2" or instrument == "NIKA":
        caldir='/home/romero/NIKA2/NIKA_SVN/Processing/Pipeline/Calibration/BP/'
        bpfile=caldir+'Transmission_2017_Jan_NIKA2_v1.fits'
        hdulist = fits.open(bpfile)

        if array == "1H":      # 1mm (260 GHz) array, Horizontal Polarization
            tbdata = hdulist[1].data # 1H
            freq = tbdata.field(0)
            tran = tbdata.field(1)
            erro = tbdata.field(2)
            atmt = tbdata.field(3)
            cfreq1h = np.sum(freq*tran)/np.sum(tran)
        
        if array == "1V":     # 1mm (260 GHz) array, Vertical Polarization
            tbdata = hdulist[2].data # 1V
            freq = tbdata.field(0)
            tran = tbdata.field(1)
            erro = tbdata.field(2)
            atmt = tbdata.field(3)
            cfreq1v = np.sum(freq*tran)/np.sum(tran)
        
        if array == "2":       # 2mm (150 GHz) array
            tbdata = hdulist[3].data # 2
            freq = tbdata.field(0)
            tran = tbdata.field(1)
            erro = tbdata.field(2)
            atmt = tbdata.field(3)
            cfreq2 = np.sum(freq*tran)/np.sum(tran)

        ### Trim the zero-frequency listing, if any.
        gi=np.where(freq > 0)
        freq = freq[gi]
        tran = tran[gi]
        erro = erro[gi]
        atmt = atmt[gi]
        
### Calculate Aperture efficiencies from information found at:
### http://www.iram.es/IRAMES/mainwiki/Iram30mEfficiencies
        Beff = 0.630         # at 210 GHz
        Aeff = Beff/1.27     # See text on webpage
        srms = (66.0*u.um).to("m")        # surface RMS (microns)
        R210 = np.exp(-4.0*np.pi*(srms/(const.c/(2.1e11*u.s**-1))).value)    #
        Gnot = Aeff/R210                   # Unphysical, but see documentation...

        Larr = const.c.value/(freq*1.0e9) # Keep calm and carry on.        
        Ruze = Gnot * np.exp(-4.0*np.pi*(srms.value)/Larr)
        NRuz = Ruze / np.max(Ruze)        # Normalize it
        band = tran * Ruze                # Bandpass, with (unnormalized) Ruze efficiency
        farr = freq

    if instrument == "SZA":

        flow = 30.0; fhigh=32.0
        farr = np.arange(flow,fhigh,0.5)  # ??? Clueless here.
        band = farr*0.0 + 1.0             # Just uniform across the range...
        
    if instrument == "ACT90":

        flow = 80.0; fhigh=110.0
        farr = np.arange(flow,fhigh,0.5)  # ??? Clueless here.
        band = farr*0.0 + 1.0             # Just uniform across the range...
        
    if instrument == "ACT150":

        flow = 130.0; fhigh=170.0
        farr = np.arange(flow,fhigh,0.5)  # ??? Clueless here.
        band = farr*0.0 + 1.0             # Just uniform across the range...
        
#########################################################################

    return band, farr

class maps:

    def __init__(self,inputs,clusters,ptsrc,bulk_comps):

        print('Reading in data from: ',inputs.fitsfile)
        #data_map, header = fits.getdata(inputs.fitsfile, header=True)
        fitshdu  = fits.open(inputs.fitsfile)
        data_map = fitshdu[0].data
        header   = fitshdu[0].header
        wt_map = fits.getdata(inputs.wtfile,inputs.wtext)
        mytab = get_xfer(inputs)
        #if np.max(wt_map) > 1.0e3:
        #    print 'It looks like your weight map is in (1/Jy)^2 or (1/K)^2'

        #import pdb;pdb.set_trace()
        
        if inputs.wtisrms == True: 
            wt_map = wt_map**(-2.0)  # I'll need to account for zeros...

        wt_map /= (inputs.rmscorr**2)

        if inputs.instrument == "ACT90" or inputs.instrument == "ACT150":
            wt_map *= 1e12
            data_map /= 1e6

        self.data   = data_map
        self.header = header
        self.wts    = wt_map


        ### I want to 'lightly' filter the wtmap... as per the Xfer function error bars.
        wtxfer      = 1.0 - (mytab[2,1:]/mytab[1,1:])**2
        wtxfer      = np.append([1],wtxfer)
        #wtmapfilt   = imf.fourier_filtering_2d(wt_map,'tab',(mytab[0,0:],wtxfer))
        wtmapfilt   = wt_map*1.0
        
        self.masked_wts = wtmapfilt.flatten()

        self.units  = inputs.units
        self.name   = inputs.name

        w = WCS(inputs.fitsfile)
        fwhm1,norm1,fwhm2,norm2,fwhm,smfw,freq,FoV = inst_params(inputs.instrument)
        myastro                                    = astrometry(header)
        
        keep        = np.array([])
        #maxclusrad  = FoV.to('arcsec').value  * 0.8
        maxclusrad  = FoV.to('arcsec').value  * 0.7
        if clusters[0].name == 'idcs' or clusters[0].name == 'HSC_3-E':
            maxclusrad  = FoV.to('arcsec').value  * 0.47

        #if bulk_comps.model == 'Beta':
        #    maxclusrad  = FoV.to('arcsec').value  * 0.4

        maxptsrcrad = fwhm.to('arcsec').value * 3.0

        MaxTheta=0

        for cluster in clusters:
            Theta500    = cluster.theta_500.to('arcsec')
            MaxTheta    = np.max([MaxTheta,Theta500.value])
            
        if maxclusrad > 2.0*MaxTheta:
            maxclusrad = 2.0*MaxTheta

        
        for cluster,geoparams in zip(clusters,bulk_comps.geoparams):
            x0,y0  = w.wcs_world2pix(cluster.ra_deg,cluster.dec_deg,0)
            x,y    = mm.get_xymap(wt_map,myastro.pixs,x0,y0)
            radmap = np.sqrt(x**2 + y**2)

            mykeep = np.array(np.where(radmap < maxclusrad)).flatten()

            if geoparams[7] > 0:
                x,y = es.rot_trans_grid(x,y,geoparams[0],geoparams[1],geoparams[2])
                x,y = es.get_ell_rads(x,y,geoparams[3],geoparams[4])                # 0.001 sec per call
                angmap = np.arctan2(y,x)
                #bi = (abs(angmap) > geoparams[7]/2.0)
                gi = (abs(angmap) <= geoparams[7]/2.0)
                toint = np.array(np.where(gi)).flatten()
                mykeep = np.intersect1d(mykeep,toint)
                # Below line is problematic.
                #geoparams[7]=0  # Set to zero so that the models populate ALL pixels (desirable for Xfer function)
                
            #rm2d   = radmap.reshape(wt_map.shape)

            print(type(mykeep),type(keep))
            print(mykeep.shape, keep.shape)
            #import pdb;pdb.set_trace()
            keep = mykeep if len(keep) == 0 else np.union1d(keep,mykeep)
            
        for ptsrc_loc in ptsrc.locs:
            ra     = ptsrc_loc[0].to('deg')
            dec    = ptsrc_loc[1].to('deg')
            x0,y0  = w.wcs_world2pix(ra.value,dec.value,0)
            x,y    = mm.get_xymap(wt_map,myastro.pixs,x0,y0)
            radmap = np.sqrt(x**2 + y**2)
            #rm2d   = radmap.reshape(wt_map.shape)

            mykeep = np.array(np.where(radmap < maxptsrcrad)).flatten()
            print(type(mykeep),type(keep))
            print(mykeep.shape, keep.shape)
            keep   = mykeep if len(keep) == 0 else np.union1d(keep,mykeep)
            #keep   = np.union1d(keep,mykeep)

        
        if inputs.instrument == "MUSTANG" or inputs.instrument == "MUSTANG2":
            #if inputs.units == 'Jy' or np.max(wt_map) < 1.0e3:
            if np.max(wt_map) < 1.0e3:
                self.wts        *= 1.0e6
                self.masked_wts *= 1.0e6

 
            #nzwts = (wt_map > 0)
            #medfact = 1.5
            #if inputs.name == 'lynx':
            #    medfact=2.0
            #keep=np.where(wt_map > np.median(wt_map[nzwts])/medfact)
            #mask=np.where(wt_map < np.median(wt_map[nzwts])/medfact)
            #print 'FYI: with a mask threshold of median(wt_map)/',medfact,', we use ',npix,\
            #'pixels, which is roughly equivalent to a circle of radius ',\
            #"{:7.1f}".format(np.sqrt(npix/np.pi)),' pixels.'

        gwts = self.masked_wts[keep]
        npix = len(gwts)
            
        print(bcolors.OKGREEN + 'FYI: with a mask around all relevant locations we use ',npix,\
            'pixels, which is roughly equivalent to a circle of radius ',\
            "{:7.1f}".format(np.sqrt(npix/np.pi)),' pixels.' + bcolors.ENDC)

        self.masked_wts      *= 0.0
        self.masked_wts[keep] = gwts
        mywts                 = self.masked_wts.reshape(wt_map.shape)
        self.masked_wts       = mywts
        #import pdb;pdb.set_trace()


def get_sz_bp_conversions(temp,instrument,bv,units='Kelvin',array="2",inter=False,beta=1.0/300.0,
                          betaz=1.0/300.0,rel=True,quiet=False,cluster='Default',RJ=False,
                          retFreqs=False,addruze=False):
    """
    + Updates: (8 March 2019), changed default units to Kelvin
    + RJ is still left as False by default, as I have a lower statement that makes it True
    + For MUSTANG-2.
    -------------------------------------------------------------------------
    This module calculates the conversion factor, f(x) or g(x), depending on whether you are
    working with Kelvin (for f(x)), or Jy/beam (for g(x)).
    
    """

    print("Beam volume: ",bv)
    
    szcv,szcu=mad.get_sz_values()
    #szcv,szcu=get_sz_values()
    freq_conv = (szcv['planck'] *1.0e9)/(szcv['boltzmann']*szcv['tcmb'])
    temp_conv = 1.0/szcv['m_e_c2']
    #bv = get_beamvolume(instrument,cluster)
#    fwhm1,norm1,fwhm2,norm2,fwhm,smfw,freq,FoV = inst_params(instrument)
    band, farr = inst_bp(instrument,array,addruze=addruze)
    #band, farr = inst_bp(instrument,array)
    fstep = np.median(farr - np.roll(farr,1))
    foo = np.where(band > 0)
    band=band[foo]
    farr=farr[foo]
############################################
    llt  = temp*temp_conv  # Lower limit on temperature (well, theta)
    ult  = temp*temp_conv  # Upper limit on temperature
    st   = temp            # Temperature (theta) step
    flow = np.min(farr)
 #   fhigh= np.ceil((np.max(farr)-flow)/fstep)*fstep+flow+fstep/100.0
    fhigh= np.max(farr)
    sx   = fstep*freq_conv
    llx  = flow*freq_conv
    nste = (fhigh-flow)/fstep
    ulx  = fhigh*freq_conv + sx/(2.0*nste)
    
#### The old way:
#    temparr,freqarr,conv = tsz.tSZ_conv_range(tlow,thigh,tstep,flow,fhigh,fstep)
#### The new way (for now):

    if inter == True:    
        tarr = np.arange(llt,ult,st)
        xarr = np.arange(llx,ulx,sx)
        T = tsz.itoh_2004_int(tarr,xarr)
    else:
        tarr, xarr, T = tsz.tSZ_conv(llt,llx,ult,ulx,st,sx)
        
    tarr, xarr, K = ksz.kSZ_conv(beta,betaz,llt,llx,ult,ulx,st,sx,rel=rel)
    
#    import pdb; pdb.set_trace()
### Let's check that the frequency spacing is close to what was given in the
### bandpass retreival:
    fq = np.abs(farr*freq_conv - xarr)/(farr*freq_conv)
    if np.max(fq) > 0.05:
        raise Exception("Frequency arrays differ significantly.")
## Else (implicit here):
    TY = T/tarr    # Divide by tarr (thetae) to get proper conversion units
    KY = K/tarr    # Divide by tarr (thetae) to get proper conversion units
    bT = np.sum(TY*band)/np.sum(band)  # Average over the bandpass
    bK = np.sum(KY*band)/np.sum(band)  # Average over the bandpass

    Fwtd = np.sum(farr*TY*band)/np.sum(TY*band)
    difa = TY-bT
    ctz  = np.where(np.abs(difa) == np.min(np.abs(difa)))
    myind = ctz[0] + np.arange(-2,3,1)
    #import pdb;pdb.set_trace()
    
    pars = np.polyfit(farr[myind],difa[myind],2)
    roots= np.roots(pars)
    rdiff= np.abs(roots - 90.0)
    myroot = roots[0] if rdiff[0] < rdiff[1] else roots[1]

    #import pdb;pdb.set_trace()
    #svarr = np.vstack((farr,band))
    #fname='/home/romero/Python/StandAlone/Comptony_Modelling/Python_M2_total_bandpass_scaled.txt'
    #np.savetxt(fname,svarr.T,fmt='%7.3f')

    JypB = tsz.Jyperbeam_factors(bv)        # Jy per beam conversion factor, from y (bT)
    xavg = np.sum(xarr*band)/np.sum(band)   # Bandpass averaged frequency; should be a reasonable approximation.
    Kepy = tsz.TBright_factors(xavg)        # Kelvin_CMB conversion factor, from y (bT)
    KCMB_per_KRJ = 1.23                     # At 90 GHz, from Tony
    #KCMB_per_KRJ = 1.286                     # At 90 GHz, with some fudge?
    if instrument == "MUSTANG2": RJ = True
    if RJ: Kepy /= KCMB_per_KRJ
    #import pdb;pdb.set_trace()
    
    tSZ_JyBeam_per_y = JypB * bT  # Just multiply by Compton y to get Delta I (tSZ)
    kSZ_JyBeam_per_t = JypB * bK  # Just multiply by tau (of electrons) to get Delta I (kSZ)
    tSZ_Kelvin_per_y = Kepy * bT  # Just multiply by Compton y to get Delta T (tSZ)
    kSZ_Kelvin_per_t = Kepy * bK  # Just multiply by tau (of electrons) to get Delta T (kSZ)

    if quiet == False:
        medfreq = np.median(farr)
        print('For a median frequency of ',medfreq,' GHz and ')
        print('Assuming a temperature of ',temp,' keV, we find the following.')
        print('The weighted (average) frequency is ',Fwtd)
        print('While the effective frequency is ',myroot)
        print('To go from Compton y to Jy/Beam, multiply by: ', tSZ_JyBeam_per_y)
        print('To go from tau to Jy/Beam (kSZ), multiply by: ', kSZ_JyBeam_per_t)
        print('To go from Compton y to Kelvin, multiply by: ', tSZ_Kelvin_per_y)
        print('To go from tau to Kelvin (kSZ), multiply by: ', kSZ_Kelvin_per_t)
        
    #if quiet == False:
    #    print(bcolors.OKGREEN + 'Assuming a temperature of ',temp,' keV, we find the following.' + bcolors.ENDC)
    #    print(bcolors.OKGREEN + 'To go from Compton y to Jy/Beam, multiply by: ', tSZ_JyBeam_per_y.value, ' '  + bcolors.ENDC)
    #    print(bcolors.OKGREEN + 'To go from tau to Jy/Beam (kSZ), multiply by: ', kSZ_JyBeam_per_t.value, ' '  + bcolors.ENDC)
    #    print(bcolors.OKGREEN + 'To go from Compton y to T_CMB, multiply by: ', tSZ_Kelvin_per_y, ' '  + bcolors.ENDC)
    #    print(bcolors.OKGREEN + 'To go from tau to T_CMB (kSZ), multiply by: ', kSZ_Kelvin_per_t, ' '  + bcolors.ENDC)
    #    print(bcolors.OKGREEN + 'To go from Compton y to T_Bright, multiply by: ', tSZ_Tbright_per_y, ' '  + bcolors.ENDC)
    #    print(bcolors.OKGREEN + 'To go from tau to T_Bright (kSZ), multiply by: ', kSZ_Tbright_per_t, ' ' + bcolors.ENDC)
    #    print(bcolors.OKGREEN + 'The inferred Jy2K value is: ',tSZ_JyBeam_per_y.value/tSZ_Tbright_per_y, ' '+ bcolors.ENDC)
    #    print(bcolors.OKGREEN + 'The bandpass average frequency (in GHz) is: ',favg, ' '  + bcolors.ENDC)
    
    
    if units == 'Kelvin':
        tSZ_return = tSZ_Kelvin_per_y; kSZ_return = kSZ_Kelvin_per_t
    else:
        tSZ_return = tSZ_JyBeam_per_y; kSZ_return = kSZ_JyBeam_per_t     

    if retFreqs:
        return tSZ_return, kSZ_return, Fwtd, myroot
    else:
        return tSZ_return, kSZ_return
        
def old_get_sz_bp_conversions(temp,instrument,bv,units='Jy/beam',array="2",inter=False,beta=1.0/300.0,
                          betaz=1.0/300.0,rel=True,quiet=False,RJ=False):

    szcv,szcu=mad.get_sz_values()
    freq_conv = (szcv['planck'] *1.0e9)/(szcv['boltzmann']*szcv['tcmb'])
    temp_conv = 1.0/szcv['m_e_c2']
    if bv == 0: bv = get_beamvolume(instrument)
    print('To calculate Jy/K for ',instrument,' we use a beam volume of ',bv ,' arcsec2')

#    fwhm1,norm1,fwhm2,norm2,fwhm,smfw,freq,FoV = inst_params(instrument)
    band, farr = inst_bp(instrument,array)
    fstep = np.median(farr - np.roll(farr,1))
    foo = np.where(band > 0)
    band=band[foo]
    farr=farr[foo]
############################################
    llt  = temp*temp_conv  # Lower limit on temperature (well, theta)
    ult  = temp*temp_conv  # Upper limit on temperature
    st   = temp            # Temperature (theta) step
    flow = np.min(farr)
 #   fhigh= np.ceil((np.max(farr)-flow)/fstep)*fstep+flow+fstep/100.0
    fhigh= np.max(farr)
    sx   = fstep*freq_conv
    llx  = flow*freq_conv
    nste = (fhigh-flow)/fstep
    ulx  = fhigh*freq_conv + sx/(2.0*nste)
    
#### The old way:
#    temparr,freqarr,conv = tsz.tSZ_conv_range(tlow,thigh,tstep,flow,fhigh,fstep)
#### The new way (for now):

    if inter == True:    
        tarr = np.arange(llt,ult,st)
        xarr = np.arange(llx,ulx,sx)
        T = tsz.itoh_2004_int(tarr,xarr)
    else:
        tarr, xarr, T = tsz.tSZ_conv(llt,llx,ult,ulx,st,sx)
        
    tarr, xarr, K = ksz.kSZ_conv(beta,betaz,llt,llx,ult,ulx,st,sx,rel=rel)
    
#    import pdb; pdb.set_trace()
### Let's check that the frequency spacing is close to what was given in the
### bandpass retreival:
    fq = np.abs(farr*freq_conv - xarr)/(farr*freq_conv)
    if np.max(fq) > 0.05:
        raise Exception("Frequency arrays differ significantly.")
## Else (implicit here):
    TY = T/tarr    # Divide by tarr (thetae) to get proper conversion units
    KY = K/tarr    # Divide by tarr (thetae) to get proper conversion units
    bT = np.sum(TY*band)/np.sum(band)  # Average over the bandpass
    bK = np.sum(KY*band)/np.sum(band)  # Average over the bandpass

    JypB = tsz.Jyperbeam_factors(bv)        # Jy per beam conversion factor, from y (bT)
    xavg = np.sum(xarr*band)/np.sum(band)   # Bandpass averaged frequency; should be a reasonable approximation.
    Kepy = tsz.TBright_factors(xavg)        # Kelvin conversion factor, from y (bT)

    #####################################################################
    KCMB_per_KRJ = 1.23                     # At 90 GHz, from Tony
    #KCMB_per_KRJ = 1.286                     # At 90 GHz, with some fudge?
    if instrument == "MUSTANG2": RJ = True
    if RJ: Kepy /= KCMB_per_KRJ
    #####################################################################

    favg = xavg/freq_conv
    
    tSZ_JyBeam_per_y  = JypB * bT  # Just multiply by Compton y to get Delta I (tSZ)
    kSZ_JyBeam_per_t  = JypB * bK  # Just multiply by tau (of electrons) to get Delta I (kSZ)
    tSZ_Kelvin_per_y  = Kepy * bT  # Just multiply by Compton y to get Delta T (tSZ)
    kSZ_Kelvin_per_t  = Kepy * bK  # Just multiply by tau (of electrons) to get Delta T (kSZ)
    Tratio            = get_tCMB_to_tRJ(favg)
    tSZ_Tbright_per_y = tSZ_Kelvin_per_y / Tratio
    kSZ_Tbright_per_t = kSZ_Kelvin_per_t / Tratio

    if quiet == False:
        print(bcolors.OKGREEN + 'Assuming a temperature of ',temp,' keV, we find the following.' + bcolors.ENDC)
        print(bcolors.OKGREEN + 'To go from Compton y to Jy/Beam, multiply by: ', tSZ_JyBeam_per_y.value, ' '  + bcolors.ENDC)
        print(bcolors.OKGREEN + 'To go from tau to Jy/Beam (kSZ), multiply by: ', kSZ_JyBeam_per_t.value, ' '  + bcolors.ENDC)
        print(bcolors.OKGREEN + 'To go from Compton y to T_CMB, multiply by: ', tSZ_Kelvin_per_y, ' '  + bcolors.ENDC)
        print(bcolors.OKGREEN + 'To go from tau to T_CMB (kSZ), multiply by: ', kSZ_Kelvin_per_t, ' '  + bcolors.ENDC)
        print(bcolors.OKGREEN + 'To go from Compton y to T_Bright, multiply by: ', tSZ_Tbright_per_y, ' '  + bcolors.ENDC)
        print(bcolors.OKGREEN + 'To go from tau to T_Bright (kSZ), multiply by: ', kSZ_Tbright_per_t, ' ' + bcolors.ENDC)
        print(bcolors.OKGREEN + 'The inferred Jy2K value is: ',tSZ_JyBeam_per_y.value/tSZ_Tbright_per_y, ' '+ bcolors.ENDC)
        print(bcolors.OKGREEN + 'The bandpass average frequency (in GHz) is: ',favg, ' '  + bcolors.ENDC)
    
    if units == 'Kelvin':
        tSZ_return = tSZ_Tbright_per_y; kSZ_return = kSZ_Tbright_per_t
        print(bcolors.OKGREEN + 'Maps are in Kelvin; using ',tSZ_return, ' '  + bcolors.ENDC)
    else:
        tSZ_return = tSZ_JyBeam_per_y.value; kSZ_return = kSZ_JyBeam_per_t.value        
        print(bcolors.OKGREEN + 'Maps are in Jy/beam; using ',tSZ_return, ' '  + bcolors.ENDC)
        
    return tSZ_return, kSZ_return

def get_beamvolume(instrument):

    ### The default is to assume a double Gaussian
    fwhm1,norm1,fwhm2,norm2,fwhm,smfw,freq,FoV = inst_params(instrument)

    sc   = 2.0 * np.sqrt(2.0*np.log(2)) # Sigma conversion (from FWHM)
    sig1 = fwhm1/sc                     # Apply the conversion
    sig2 = fwhm2/sc                     # Apply the conversion
    bv1  = 2.0*np.pi * norm1*sig1**2    # Calculate the integral
    bv2  = 2.0*np.pi * norm2*sig2**2    # Calculate the integral
    beamvolume = bv1 + bv2  # In units of FWHM**2 (should be arcsec squared) 

    print('Using ',beamvolume,' for ',instrument)
    
    return beamvolume

def get_sz_conversion(temp,instrument,beta=0.0,betaz=0.0):
    """
    Returns the tSZ and kSZ conversions to Jy/beam for a given instrument,
    for a given electron temperature, velocity relative to the speed of
    light (beta), and beta along the line of sight (betaz).
    
    Parameters
    __________
    instrument : MUSTANG, BOLOCAM, or NIKA
    target: The target of the cluster
    real: True or False (do you want to analyze/fit real data or
             'virtual' (simulated) data?).import numpy as np

    Returns
    -------
    CtSZ, CkSZ
    """

    bv = get_beamvolume(instrument)
    fwhm1,norm1,fwhm2,norm2,fwhm,smfw,freq,FoV = inst_params(instrument)
 #   conv = tsz.tSZ_conv_single(temp,freq.value)
    bpsr = tsz.intensity_factors(freq.value,bv)
    szcv,szcu=mad.get_sz_values()
    freq_conv = (szcv['planck'] *1.0e9)/(szcv['boltzmann']*szcv['tcmb'])
    temp_conv = 1.0/szcv['m_e_c2']
    conv = tsz.itoh_2004_int(temp*temp_conv,freq*freq_conv)

    print(conv,bpsr)
    yJyBeam = conv.item(0)*bpsr/(temp*temp_conv) # Conversion from Compton y to Jy/beam

    return yJyBeam

###################################################################
### Under verification below.

def astro_from_hdr(hdr):
    
    xsz = hdr['naxis1']
    ysz = hdr['naxis2']
    xar = np.outer(np.arange(xsz),np.zeros(ysz)+1.0)
    yar = np.outer(np.zeros(xsz)+1.0,np.arange(ysz))
    ####################

    w = wcs.WCS(hdr)
    #import pdb;pdb.set_trace()
    
    xcen = hdr['CRPIX1']
    ycen = hdr['CRPIX2']
    dxa = xar - xcen
    dya = yar - ycen
    ### RA and DEC in degrees:
    if 'CDELT1' in hdr.keys():
        ras = dxa*hdr['CDELT1'] + hdr['CRVAL1']
        decs= dya*hdr['CDELT2'] + hdr['CRVAL2']
        pixs= abs(hdr['CDELT1'] * hdr['CDELT2'])**0.5 * 3600.0    
    if 'CD1_1' in hdr.keys():
        ras = dxa*hdr['CD1_1'] + dya*hdr['CD2_1'] + hdr['CRVAL1']
        decs= dxa*hdr['CD1_2'] + dya*hdr['CD2_2'] + hdr['CRVAL2']
        pixs= abs(hdr['CD1_1'] * hdr['CD2_2'])**0.5 * 3600.0
    if 'PC1_1' in hdr.keys():
        pcmat = w.wcs.get_pc()
        ras = dxa*pcmat[0,0]*hdr['CDELT1'] + \
              dya*pcmat[1,0]*hdr['CDELT2'] + hdr['CRVAL1']
        decs= dxa*pcmat[0,1]*hdr['CDELT1'] + \
              dya*pcmat[1,1]*hdr['CDELT2'] + hdr['CRVAL2']
        pixs= abs(pcmat[0,0]*hdr['CDELT1'] * \
                  pcmat[1,1]*hdr['CDELT2'])**0.5 * 3600.0

    pixs = pixs*u.arcsec
    ### Do I want to make ras and decs Angle objects??
    ras  = ras*u.deg; decs = decs*u.deg 
    
    return ras, decs, pixs

def get_astro(file):

    hdu = fits.open(file)
    hdr = hdu[0].header
    image_data = hdu[0].data

    ras, decs, pixs = astro_from_hdr(hdr)

    return image_data, ras, decs, hdr, pixs

class astrometry:

    def __init__(self,hdr):

        ras,decs,pixs = astro_from_hdr(hdr)
        self.ras = ras
        self.decs= decs
        self.pixs= pixs
        
def get_xfer(inputs):

    if inputs.tabformat == 'ascii':
        tab = np.loadtxt(inputs.tabfile, comments=inputs.tabcomments)
    if inputs.tabformat == 'fits':
        ktab = fits.getdata(inputs.tabfile)
        xtab = fits.getdata(inputs.tabfile,ext=1)
        tab = np.vstack((ktab,xtab))
        
    if inputs.tabdims == '1D':
        if inputs.instrument == "MUSTANG" or inputs.instrument == "MUSTANG2":
            tab = tab.T            # Transpose the table.
#        import pdb;pdb.set_trace()
        if inputs.tabextend==True:
            tdim = tab.shape
            #import pdb;pdb.set_trace()
            pfit = np.polyfit(tab[0,tdim[1]//2:],tab[1,tdim[1]//2:],1)
            addt = np.max(tab[0,:]) * np.array([2.0,4.0,8.0,16.0,32.0])
            extt = np.polyval(pfit,addt)
            ### For better backwards compatability I've editted to np.vstack instead of np.stack
            if tdim[0] == 2:
                foo = np.vstack((addt,extt)) # Mar 5, 2018
            else:
                pfit2 = np.polyfit(tab[0,tdim[1]//2:],tab[2,tdim[1]//2:],1)
                extt2 = np.polyval(pfit2,addt)
                foo = np.vstack((addt,extt,extt2)) # Mar 5, 2018

            print(tab.shape, foo.shape)
            tab = np.concatenate((tab,foo),axis=1)
#            newt = [np.append(tab[0,:],addt),np.append(tab[1,:],extt)]
            
    return tab

############################################################################

def get_conv_factor(instrument,bv=None):
    
    fwhm1,norm1,fwhm2,norm2,fwhm,smfw,freq,FoV = inst_params(instrument)
    #freq = 90.0 * u.GHz;     instrument='MUSTANG2'
    szcv,szcu = mad.get_sz_values()
    x = szcv["planck"]*(freq.to("Hz")).value / (szcv["boltzmann"]*szcv["tcmb"])
    if bv is None:
        bv = get_beamvolume(instrument)

    fofx = x * (np.exp(x) + 1.0)/(np.exp(x) - 1.0) - 4.0 # Delta T / T * y
    gofx = fofx * x**4 * np.exp(x) / (np.exp(x) - 1)**2  # Delta I / I * y

    B_nu = 2.0*((const.h*freq**3)/(const.c**2 * u.sr)).decompose()
    B_nu *= 1.0/(np.exp(x) - 1.0)
    JyperSrK = (B_nu/(szcv["tcmb"]*u.K)).to("Jy / (K sr)")
    JyperAsK = JyperSrK.to("Jy / (K arcsec2)")
    
    ### This value assumes true black-body spectrum:
    JyperK_SZ = JyperAsK*bv*gofx/fofx
    print('Your "naive" value for Jy per Kelvin_CMB is: ',JyperK_SZ)
    
    ### Radiance per Kelvin:
    I_per_K = 2.0*freq**2 * const.k_B / (const.c**2 * u.sr)
    IpK = I_per_K.to("Hz J s / (K m2 sr)")
    WperSr = IpK.to("W / (Hz K m2 sr)")
    JyperSr = WperSr.to("Jy / (K sr)")
    Jy_per_K_as = JyperSr.to("Jy / (K arcsec2)")

    ###  And this one assumes the RJ law:
    JyperK_RJ = Jy_per_K_as*bv
    print('Your "naive" value for Jy per Kelvin_RJ is: ',JyperK_RJ)

    return JyperK_RJ

############################################################################

def get_bv_from_Jy2K(Jy2K,instrument):

    #fwhm1,norm1,fwhm2,norm2,fwhm,smfw,freq,FoV = inst_params(instrument)
    freq = 9e10 * u.Hz
    wl = (const.c / freq).to('m').value
    a2r = ((1.0/3600.0/180.)*np.pi)**2
    rjk = (1e-3 *wl**2)/(2.0*1.38064)  # 1e-26 / 1e-23 = 1e-3 (1e-26 for Jy to W; 1e-23 for OoM of Boltzmann constant)
    vol = Jy2K / (a2r/rjk) * u.arcsec**2

    return vol

def get_tCMB_to_tRJ(myfreq):

    ### Assume freq to be in GHz and a value (not a unit/quantity)
    szcv,szcu=mad.get_sz_values()
    freq_conv = (szcv['planck'] *1.0e9)/(szcv['boltzmann']*szcv['tcmb'])
    x = myfreq*freq_conv

    numer = (2.0*szcv['boltzmann']*szcv['tcmb']*u.keV/u.sr)*(np.exp(x)-1.0)**2 *(myfreq*u.GHz)**2
    denom = (szcu["Icmb"] * x**4 * np.exp(x) * (const.c**2))
    Tratio = (numer/denom).decompose()

    #fts = 2.0*spconst.value("Planck constant")*u.J*u.s*(myfreq*u.GHz)**3 / (const.c**2)
    #fts = fts.to('Jy')
    #ICMB = fts * (1.0 / (np.exp(x)-1.0))
    
    #ICMB = szcu['Jycmb'] * x**3 / (np.exp(x)-1.0)
    #myboltz = spconst.value("Boltzmann constant") * u.J
    #IRJ  = (myboltz*szcv['tcmb'])*2.0*(myfreq*u.GHz)**2 / (const.c**2)
    #IRJ  = IRJ.to("Jy") / u.sr

    return Tratio
